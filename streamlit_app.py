# -*- coding: utf-8 -*-
"""
RE_price (웹앱) — 서울 빌라(연립다세대) 호별 예상 거래가 산출기
실거래가 = 토지+건물. 대지평당가(N년) = 토지하한 + (신축총 − 토지하한) × 국세청 공식 잔가율.
지번 조회 시 그 위 건축물의 호별 대지권비율·평당가·예상가(층별 보정)를 함께 산출.
같은 폴더: prices.json, PNU_coords.npz, dongnames.csv, ho.json
"""
import os, csv, glob, json, io
import numpy as np
import streamlit as st
from openpyxl import Workbook

PYEONG = 3.305785
IDW_K = 8; IDW_P = 2.0; MIN_DONG_PTS = 3
HERE = os.path.dirname(os.path.abspath(__file__))
DEF_B = {"beta": 500.0, "floor": 0.10, "rc": {"rate": 0.018, "life": 50}, "brick": {"rate": 0.0225, "life": 40}}
# 층별 보정 (동 대비): 반지하 0.625 · 1층 0.854 · 2~3층 1.0(기준) · 4층↑ 1.093
FLOOR_MULT = {"B": 0.625, "1": 0.854, "2": 1.0, "4": 1.093}
FLOOR_LABEL = {"B": "반지하", "1": "1층", "2": "2~3층", "4": "4층↑"}


def _load(here):
    cz = glob.glob(os.path.join(here, "PNU_coords*.npz")) or glob.glob(os.path.join(here, "*.npz"))
    z = np.load(cz[0], allow_pickle=False)
    pnu = z["pnu"].astype("U19")
    x = z["x"].astype(np.float64); y = z["y"].astype(np.float64)
    area = z["area"].astype(np.float64) if "area" in z.files else np.full(len(pnu), np.nan)
    age = z["age"].astype(np.int16) if "age" in z.files else np.full(len(pnu), -1, np.int16)
    struct = z["struct"].astype(np.uint8) if "struct" in z.files else np.zeros(len(pnu), np.uint8)
    o = np.argsort(pnu); pnu, x, y, area, age, struct = pnu[o], x[o], y[o], area[o], age[o], struct[o]
    prices = json.load(open(os.path.join(here, "prices.json"), encoding="utf-8"))
    # 호별 — 경량(npz) 우선
    ho = {}
    if os.path.exists(os.path.join(here, "ho_lite.npz")):
        hz = np.load(os.path.join(here, "ho_lite.npz"), allow_pickle=False)
        for pp, hm, fl, ar in zip(hz["pnu"].astype("U19"), hz["ho"], hz["fl"], hz["area"]):
            ho.setdefault(str(pp), []).append([str(hm), str(fl), float(ar)])
    elif os.path.exists(os.path.join(here, "ho.json")):
        ho = json.load(open(os.path.join(here, "ho.json"), encoding="utf-8"))
    # 도로명 — 경량(npz) 우선
    road = {"keys": None, "pnu": None, "d": None}
    if os.path.exists(os.path.join(here, "road_lite.npz")):
        rz = np.load(os.path.join(here, "road_lite.npz"), allow_pickle=False)
        road = {"keys": rz["keys"], "pnu": rz["pnu"].astype("U19"), "d": None}
    elif os.path.exists(os.path.join(here, "road.json")):
        road = {"keys": None, "pnu": None, "d": json.load(open(os.path.join(here, "road.json"), encoding="utf-8"))}
    dn = {}
    with open(os.path.join(here, "dongnames.csv"), encoding="utf-8-sig", newline="") as f:
        for row in csv.DictReader(f):
            c = (row.get("동코드") or "").strip(); n = (row.get("동명") or "").strip()
            if c and n: dn[c] = n
    kxx, kxy, kv, kdong, kpnu = [], [], [], [], []
    for p, pair in prices["parcel"].items():
        i = np.searchsorted(pnu, p)
        if i < len(pnu) and pnu[i] == p:
            kxx.append(x[i]); kxy.append(y[i]); kv.append(pair[:2]); kdong.append(p[:10]); kpnu.append(p)
    kxx = np.asarray(kxx, float); kxy = np.asarray(kxy, float); kv = np.asarray(kv, float)
    kpnu = np.asarray(kpnu, dtype="U19")
    dong_idx = {}
    for idx, dgc in enumerate(kdong):
        dong_idx.setdefault(dgc, []).append(idx)
    for dgc in dong_idx:
        dong_idx[dgc] = np.asarray(dong_idx[dgc], int)
    # 필지별 실거래 (근거거래 표시용) — txn_lite.npz 다건
    tx = {}
    if os.path.exists(os.path.join(here, "txn_lite.npz")):
        tz = np.load(os.path.join(here, "txn_lite.npz"), allow_pickle=False)
        for pp, py_, amt, la, fl, ag, y6 in zip(tz["pnu"].astype("U19"), tz["py"], tz["amt"],
                                                tz["land"], tz["fl"], tz["age"], tz["ym"]):
            tx.setdefault(str(pp), []).append({"평당": int(py_), "금액": int(amt), "대지": float(la),
                                               "층": int(fl), "연식": int(ag), "시점": int(y6)})
    return {"pnu": pnu, "x": x, "y": y, "area": area, "age": age, "struct": struct,
            "prices": prices, "ho": ho, "road": road, "tx": tx,
            "dn": dn, "rev": {n: c for c, n in dn.items()},
            "kxx": kxx, "kxy": kxy, "kv": kv, "kpnu": kpnu, "dong_idx": dong_idx,
            "building": prices.get("building", DEF_B)}


def road_get(road, key):
    if road.get("d") is not None:
        return road["d"].get(key)
    keys = road.get("keys")
    if keys is None or len(keys) == 0:
        return None
    i = int(np.searchsorted(keys, key))
    if 0 <= i < len(keys) and keys[i] == key:
        return str(road["pnu"][i])
    return None


@st.cache_resource(show_spinner="데이터 불러오는 중 ...")
def load_data():
    return _load(HERE)


def normalize_pnu(s):
    return "".join(ch for ch in str(s).strip().replace("-", "").replace(" ", "") if ch.isdigit())


def jibun_from_pnu(p, dn):
    bon, bu = int(p[11:15]), int(p[15:19])
    번지 = f"{bon}" if bu == 0 else f"{bon}-{bu}"
    동 = dn.get(p[:10], "")
    return (동 + " " + 번지).strip() if 동 else 번지


def address_to_pnu(addr, rev):
    if not addr: return None
    toks = [t for t in addr.replace(",", " ").replace("\t", " ").split()
            if t not in ("서울특별시", "서울시", "서울")]
    if len(toks) < 2: return None
    번지 = toks[-1]; 동 = " ".join(toks[:-1])
    code = rev.get(동)
    if not code: return None
    san = "2" if 번지.startswith("산") else "1"
    번지 = 번지.lstrip("산")
    bon, bu = (번지.split("-") + ["0"])[:2] if "-" in 번지 else (번지, "0")
    try:
        bon, bu = int(bon), int(bu)
    except ValueError:
        return None
    return f"{code}{san}{bon:04d}{bu:04d}"


def norm_road(s):
    import re
    s = re.sub(r"\(.*?\)", "", str(s).strip())
    for k in ("서울특별시", "서울시", "서울"):
        if s.startswith(k):
            s = s[len(k):]
    return re.sub(r"\s+", " ", s).strip()


def resolve_address(addr, D):
    if not addr:
        return None
    key = norm_road(addr)
    hit = road_get(D.get("road", {}), key)
    if hit:
        return hit
    return address_to_pnu(addr, D["rev"])


def residual_rate(age, structure, b):
    grp = b.get("brick" if structure == "brick" else "rc", {"rate": 0.018})
    return max(b.get("floor", 0.10), 1.0 - grp["rate"] * max(0, age))


def _idw_in_dong(D, qx, qy, dong_code):
    cand = D["dong_idx"].get(dong_code)
    if cand is None or len(cand) < MIN_DONG_PTS:
        return None
    dx = D["kxx"][cand] - qx; dy = D["kxy"][cand] - qy
    d2 = dx * dx + dy * dy
    k = min(IDW_K, len(d2))
    sel = np.argpartition(d2, k - 1)[:k] if k < len(d2) else np.arange(len(d2))
    d = np.sqrt(d2[sel]); o = np.argsort(d); d = d[o]; sel = sel[o]
    gsel = cand[sel]                                  # 전역 anchor 인덱스
    vals = D["kv"][gsel]
    if d[0] == 0:
        w = None
        wv = vals[0]
    else:
        w = 1.0 / (d ** IDW_P)
        wv = (w[:, None] * vals).sum(0) / w.sum()
    nb = [{"pnu": str(D["kpnu"][gsel[j]]), "dist": float(d[j])} for j in range(len(gsel))]
    return wv.tolist(), float(d[0]), nb


def estimate_one(pnu, D):
    P = D["prices"]
    p = normalize_pnu(pnu)
    if len(p) != 19:
        return {"PNU": pnu, "지번": "", "토지면적_㎡": None, "토지면적_평": None,
                "신축총_평당_만원": None, "토지하한_평당_만원": None, "추정방식": "오류(19자리 아님)", "참고": "", "근거": None}
    지번 = jibun_from_pnu(p, D["dn"])
    i = np.searchsorted(D["pnu"], p)
    has = i < len(D["pnu"]) and D["pnu"][i] == p
    면적 = round(float(D["area"][i]), 1) if has and not np.isnan(D["area"][i]) else None
    면적평 = round(면적 / PYEONG, 1) if 면적 is not None else None
    det_age = int(D["age"][i]) if has else -1
    det_struct = {1: "rc", 2: "brick"}.get(int(D["struct"][i]) if has else 0)

    def out(pair, 방식, 참고, 근거=None):
        return {"PNU": p, "지번": 지번, "토지면적_㎡": 면적, "토지면적_평": 면적평,
                "신축총_평당_만원": round(float(pair[0]), 1), "토지하한_평당_만원": round(float(pair[1]), 1),
                "건물_사용연수": det_age, "건물_구조": det_struct, "추정방식": 방식, "참고": 참고, "근거": 근거}

    if p in P["parcel"]:
        return out(P["parcel"][p][:2], "실측", "실거래", [{"pnu": p, "dist": 0.0}])
    if has:
        r = _idw_in_dong(D, D["x"][i], D["y"][i], p[:10])
        if r is not None:
            pair, nd, nb = r
            return out(pair, "공간보간(동내)", f"최근접 {nd:.0f}m", nb)
    if p[:10] in P["dong"]:
        rec = P["dong"][p[:10]]
        return out(rec[:2], "동 추정", f"동 {rec[2] if len(rec) > 2 else 0}건")
    if p[:5] in P["gu"]:
        rec = P["gu"][p[:5]]
        return out(rec[:2], "구 추정", f"구 {rec[2] if len(rec) > 2 else 0}건")
    return out(P["total_med"], "전체 추정", "")


def units_for(D, pnu, base_pyeong):
    """호별 [호명, 층, 보정, 대지권㎡, 대지권평, 호평당가, 예상가] — base_pyeong=대지평당가(N년)"""
    rows = []
    for 호명, fc, 지분 in D["ho"].get(pnu, []):
        m = FLOOR_MULT.get(fc, 1.0)
        호평당 = base_pyeong * m
        대지평 = 지분 / PYEONG
        rows.append({"호": 호명, "층": FLOOR_LABEL.get(fc, fc), "층보정": m,
                     "대지권_㎡": round(지분, 2), "대지권_평": round(대지평, 2),
                     "호_대지평당가_만원": round(호평당, 1),
                     "호_예상가_만원": round(호평당 * 대지평)})
    # 층 → 호명 순 정렬
    order = {"반지하": 0, "1층": 1, "2~3층": 2, "4층↑": 3}
    rows.sort(key=lambda r: (order.get(r["층"], 9), r["호"]))
    return rows


def won(manwon):
    if manwon is None: return "-"
    eok = manwon / 10000.0
    return f"{manwon:,.0f}만원 ({eok:,.1f}억)" if eok >= 1 else f"{manwon:,.0f}만원"


def units_xlsx(지번, rows):
    wb = Workbook(); ws = wb.active; ws.title = "호별"
    cols = ["호", "층", "층보정", "대지권_㎡", "대지권_평", "호_대지평당가_만원", "호_예상가_만원"]
    ws.append([지번] + [""] * (len(cols) - 1)); ws.append(cols)
    for r in rows:
        ws.append([r.get(c) for c in cols])
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf.getvalue()


st.set_page_config(page_title="RE_price 호별 예상가", page_icon="🏠", layout="centered")
st.title("🏠 서울 빌라(연립다세대) 호별 예상 거래가 산출기")
st.caption("실거래가 = 토지+건물. 대지평당가(N년) = 토지하한 + (신축총 − 토지하한) × 국세청 공식 잔가율. "
           "지번을 조회하면 그 위 건축물의 호별 대지권비율·평당가·예상가를 층별 보정해 산출합니다.")

D = load_data()
B = D.get("building", DEF_B)
_samp = next(iter(D["prices"]["parcel"].values()), [0, 0])
if not isinstance(_samp, (list, tuple)) or len(_samp) < 2:
    st.error("⚠️ prices.json이 옛 버전입니다. 새 streamlit_app.py·prices.json·ho.json을 함께 올린 뒤 Reboot 하세요.")
    st.stop()
st.success(f"준비 완료 — 실측(빌라) {len(D['prices']['parcel']):,}지번 · 호데이터 {len(D['ho']):,}필지")
st.warning("⚠️ 참고용 추정치이며 평균 ±20% 오차가 있을 수 있습니다. 실제 거래·감정 시 인근 실거래로 반드시 교차 확인하세요.")

st.markdown("**감가 조건**")
auto = st.checkbox("건축물대장에서 사용연수·구조 자동 적용", value=True)
if auto:
    man_age, struct_manual = None, None
    st.caption("조회한 지번에 건축물이 있으면 사용연수·구조를 자동 반영합니다. 정보가 없으면 예상평당가(신축 기준)만 표시됩니다.")
else:
    cc1, cc2 = st.columns(2)
    man_age = cc1.number_input("사용연수(년)", min_value=0, max_value=60, value=20, step=1)
    struct_manual = "brick" if cc2.radio("구조", ["철근콘크리트", "벽돌·연와"], horizontal=True) == "벽돌·연와" else "rc"

tab1, tab2, tab3 = st.tabs(["① 주소 조회 (지번/도로명)", "② PNU 조회", "③ 여러 개(엑셀, 2시트)"])


def resolve_age(r):
    if man_age is not None:
        return man_age, struct_manual, "수동"
    if r.get("건물_사용연수", -1) >= 0:
        return int(r["건물_사용연수"]), (r.get("건물_구조") or "rc"), "자동"
    return None, None, None


def parcel_pyeong(r):
    """대지평당가(자동/수동 사용연수 반영) 계산. 반환 (평당, 사용연수, 구조명)"""
    T0 = r["신축총_평당_만원"]; TL = r["토지하한_평당_만원"]
    age, sk, src = resolve_age(r)
    a = age if age is not None else 0
    sk = sk or "rc"
    평당 = TL + (T0 - TL) * residual_rate(a, sk, B)
    구조명 = "벽돌·연와" if sk == "brick" else "철근콘크리트"
    return round(평당, 1), (age if age is not None else None), 구조명


def show(r):
    if r.get("신축총_평당_만원") is None:
        st.warning(r["추정방식"]); return
    T0 = r["신축총_평당_만원"]; TL = r["토지하한_평당_만원"]
    지번 = r["지번"] or r["PNU"]
    면적 = (f"{r['토지면적_㎡']:,.1f}㎡ ({r['토지면적_평']:,.1f}평)"
          if r["토지면적_㎡"] is not None else "면적정보없음")
    age, struct_key, src = resolve_age(r)
    has_ho = r["PNU"] in D["ho"]

    if age is None and not has_ho:
        st.metric(f"{지번} · 예상 평당가 (신축 기준)", f"{T0:,.0f} 만원/평", help="건축물 정보 없음")
        st.write(f"필지 {면적} · [{r['추정방식']} {r['참고']}] · 건축물 정보 없음")
        return

    a = age if age is not None else 0
    sk = struct_key or "rc"
    평당 = TL + (T0 - TL) * residual_rate(a, sk, B)
    label = 지번 + " · 대지평당가"
    if age is not None:
        구조명 = "벽돌·연와" if sk == "brick" else "철근콘크리트"
        label += f"  ({구조명}·사용 {a}년·{src})"
    st.metric(label, f"{평당:,.0f} 만원/평")
    st.write(f"신축 0년 상한 {T0:,.0f} 만원/평 · 필지 {면적} · [{r['추정방식']} {r['참고']}]")

    rows = units_for(D, r["PNU"], 평당)
    if rows:
        st.markdown(f"**호별 예상 ({len(rows)}호)**")
        view = [{"호": u["호"], "대지권_㎡": u["대지권_㎡"], "대지권_평": u["대지권_평"],
                 "호_대지평당가_만원": u["호_대지평당가_만원"], "호_예상가_만원": u["호_예상가_만원"]} for u in rows]
        st.dataframe(view, use_container_width=True, hide_index=True)
        st.download_button("⬇️ 호별 엑셀(.xlsx)", data=units_xlsx(지번, rows),
                           file_name=f"RE_price_호별_{r['PNU']}.xlsx",
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    # ▣ 추정 근거 실거래 — 지번별 추정가와 비교
    basis = r.get("근거")
    if basis:
        rows_b = []
        for b in basis:
            for t in D["tx"].get(b["pnu"], []):
                ym = t.get("시점", 0)
                rows_b.append({
                    "지번": jibun_from_pnu(b["pnu"], D["dn"]),
                    "거리": "실측" if b["dist"] == 0 else f"{round(b['dist']):,}m",
                    "_d": b["dist"],
                    "시점": f"{(ym//100)%100:02d}.{ym%100:02d}" if ym else "-",
                    "_ym": ym,
                    "실거래_평당_만원": t.get("평당", 0),
                    "거래가_만원": t.get("금액", 0),
                    "층": t.get("층", 0), "연식": t.get("연식", 0)})
        rows_b.sort(key=lambda x: (x["_d"], -x["_ym"]))
        rows_b = rows_b[:12]
        if rows_b:
            st.markdown(f"**추정 근거 실거래 — 추정 대지평당가 `{평당:,.0f}만원/평`과 비교**")
            view_b = [{k: v for k, v in row.items() if not k.startswith("_")} for row in rows_b]
            st.dataframe(view_b, use_container_width=True, hide_index=True)

    st.caption("⚠️ 참고용 추정치이며 평균 ±20% 오차가 있을 수 있습니다.")


def batch_xlsx(rows):
    """2시트: 평당가 리스트 + 호별 리스트"""
    wb = Workbook()
    ws1 = wb.active; ws1.title = "평당가"
    c1 = ["PNU", "지번", "토지면적_㎡", "토지면적_평", "사용연수", "구조",
          "대지평당가_만원", "예상거래가_만원", "신축총_평당_만원", "토지하한_평당_만원", "추정방식"]
    ws1.append(c1)
    ws2 = wb.create_sheet("호별")
    c2 = ["지번", "PNU", "호", "층", "대지권_㎡", "대지권_평", "호_대지평당가_만원", "호_예상가_만원", "사용연수", "구조"]
    ws2.append(c2)
    for r in rows:
        ws1.append([r.get(c) for c in c1])
        base = r.get("대지평당가_만원")
        if base is None or r["PNU"] not in D["ho"]:
            continue
        for u in units_for(D, r["PNU"], base):
            ws2.append([r.get("지번"), r["PNU"], u["호"], u["층"], u["대지권_㎡"], u["대지권_평"],
                        u["호_대지평당가_만원"], u["호_예상가_만원"], r.get("사용연수"), r.get("구조")])
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf.getvalue()


with tab1:
    addr = st.text_input("주소 (지번 또는 도로명)", placeholder="예: 영등포구 문래동6가 29  /  영등포구 문래로 191")
    if st.button("조회", key="b1") and addr.strip():
        p = resolve_address(addr, D)
        if not p:
            st.warning("주소를 PNU로 못 바꿨습니다. 지번('자치구 동 번지') 또는 도로명('자치구 도로명 번호')을 확인하세요.")
        else:
            show(estimate_one(p, D))

with tab2:
    pnu = st.text_input("PNU(19자리)", placeholder="예: 1168010100107810028")
    if st.button("조회", key="b2") and pnu.strip():
        show(estimate_one(pnu, D))

with tab3:
    st.write("주소(지번/도로명) 또는 PNU를 줄마다 하나씩 붙여넣으세요. 엑셀은 2시트(평당가·호별)로 저장됩니다.")
    text = st.text_area("목록", height=160, placeholder="강남구 역삼동 781-28\n영등포구 문래로 191\n1168010100107810028")
    if st.button("조회", key="b3"):
        items = [t.strip() for t in text.splitlines() if t.strip()]
        if not items:
            st.info("주소를 붙여넣으세요.")
        else:
            rows = []
            for it in items:
                pn = it if (it.isdigit() and len(normalize_pnu(it)) == 19) else resolve_address(it, D)
                if not pn:
                    continue
                r = estimate_one(pn, D)
                if r.get("신축총_평당_만원") is not None:
                    평당, ag, 구조명 = parcel_pyeong(r)
                    면적평 = r.get("토지면적_평")
                    r["사용연수"] = ag if ag is not None else 0
                    r["구조"] = 구조명
                    r["대지평당가_만원"] = 평당
                    r["예상거래가_만원"] = round(평당 * 면적평) if 면적평 is not None else None
                rows.append(r)
            if not rows:
                st.warning("유효한 PNU/주소가 없습니다.")
            else:
                n_ho = sum(len(D["ho"].get(r["PNU"], [])) for r in rows if r.get("대지평당가_만원") is not None)
                st.success(f"평당가 {len(rows)}건 · 호별 {n_ho}호")
                st.dataframe([{k: r.get(k) for k in ["PNU", "지번", "사용연수", "구조", "대지평당가_만원", "예상거래가_만원"]} for r in rows],
                             use_container_width=True, hide_index=True)
                st.download_button("⬇️ 엑셀 2시트(평당가+호별) 저장", data=batch_xlsx(rows),
                                   file_name="RE_price_대량조회.xlsx",
                                   mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

st.divider()
st.caption("건물원가 500만원/㎡ × 국세청 공식 잔가율(철근콘크리트 50년·1.8%, 벽돌·연와 40년·2.25%, 최종잔존 10%). "
           "직거래 제외 · 법정동 경계 공간보간 · 호별 대지권 AL_D006(2026.06).")
